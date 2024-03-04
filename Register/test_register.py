from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest

class Test_tobeto_giris():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    
    def teardown_method(self): 
        self.driver.quit()
    
    def getData():
        excel= openpyxl.load_workbook("data/register.xlsx")
        sheet= excel["successful_register"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            AD=sheet.cell(i,1).value
            SOYAD= sheet.cell(i,2).value
            EPOSTA= sheet.cell(i,3).value
            ŞİFRE= sheet.cell(i,4).value
            ŞİFRE_TEKRARI=sheet.cell(i,5).value
            TEL= sheet.cell(i,6).value
            data.append((AD,SOYAD,EPOSTA,ŞİFRE,ŞİFRE_TEKRARI,TEL))
        return data
    
    #Başarılı kayıt olma işlemi
    @pytest.mark.parametrize("AD,SOYAD,EPOSTA,ŞİFRE,ŞİFRE_TEKRARI,TEL", getData())
    def test_successfulregister(self,AD,SOYAD,EPOSTA,ŞİFRE,ŞİFRE_TEKRARI,TEL):
        kayit_ol_button= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_OL)))
        kayit_ol_button.click()
        sleep(5)

        adinput= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_NAME)))
        adinput.send_keys(AD)
        sleep(5)

        soyadinput= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_LASTNAME)))
        soyadinput.send_keys(SOYAD)
        sleep(3)

        eposta= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_EPOSTA)))
        eposta.send_keys(EPOSTA)
        sleep(3)

        sifreinput= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_SİFRE)))
        sifreinput.send_keys(ŞİFRE)
        sleep(3)

        againinput= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.KAYIT_SİFRE_AGAİN)))
        againinput.send_keys(ŞİFRE_TEKRARI)
        sleep(3)

        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)

        girisbutton = self.driver.find_element(By.XPATH, "//*[@id='__next']/div/main/section/div/div/div[1]/div/div/button")
        girisbutton.click()
        sleep(5) 
        
        acik_riza_metni = self.driver.find_element(By.XPATH, gC.ACIK_RIZA_METNİ)
        acik_riza_metni.click()
        sleep(4) 

        uyelik_sozlesmesi = self.driver.find_element(By.XPATH, gC.UYELIK_SOZLESMESİ)
        uyelik_sozlesmesi.click()
        sleep(4)

        eposta_izni= self.driver.find_element(By.XPATH, gC.EPOSTA_İZİN)
        eposta_izni.click()
        sleep(4)

        arama_izni= self.driver.find_element(By.XPATH, gC.ARAMA_İZİN)
        arama_izni.click()
        sleep(4)

        phoneinput= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.PHONEE)))
        phoneinput.send_keys(TEL)
        sleep(5)



        robot= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, "/html/body/div[4]/div/div/div/div/div/div[2]/div/div/div/iframe")))
        robot.click()
        sleep(4)
        
        iframe=self.driver.find_element(By.XPATH, gC.IFRAME )
        self.driver.switch_to.frame(iframe)
        sleep(1)

        captcha=self.driver.find_element(By.XPATH,"//*[@id='recaptcha-anchor']")
        captcha.click()
        sleep(10)

        self.driver.switch_to.default_content()
        sleep(10)
        
        continueButton = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, "/html/body/div[4]/div/div/div/div/div/div[3]/button[2]")))
        continueButton.click()
        sleep(5) 

        self.driver.execute_script("window.scrollTo(0,0)")
        sleep(3)
        
        # Mesaj kutusunu bulma ve içeriğini alıp kontrol etme
        message_box_xpath = "//*[@id='__next']/div/main/section/div/div/div/div/span"
        message_element = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, message_box_xpath)))
        message_text = message_element.text

        message=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='__next']/div/main/section/div/div/div/div/span")))
        assert message.text== "Tobeto Platform'a kaydınız başarıyla gerçekleşti.\nGiriş yapabilmek için e-posta adresinize iletilen doğrulama linkine tıklayarak hesabınızı aktifleştirin." 

    

        
