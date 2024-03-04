from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest

class Test_tobeto_giris():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    
    def teardown_method(self): 
        self.driver.quit()
    def getData():
        excel= openpyxl.load_workbook("data/data.xlsx")
        sheet= excel["profile"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            email=sheet.cell(i,1).value
            password= sheet.cell(i,2).value
            lessthanfive = sheet.cell(i,3).value
            morethanfifty = sheet.cell(i,4).value
            startcalendarchoice =sheet.cell(i,5).value
            finishStartcalendarchoice =sheet.cell(i,6).value
            finishEndcalendarchoice =sheet.cell(i,7).value
            morethanthreehundered =sheet.cell(i,8).value
            data.append((email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered))
        return data
     
#TEST CASE 1: “Deneyimlerim” Bölümündeki Alanların Listelenmesi
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_Listing_of_Fields_in_the_My_Experiences_Section(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)


        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        experiences_page_title= self.driver.find_element(By.CLASS_NAME, gC.KURUM_ADI)
        experiences_page_title1= self.driver.find_element(By.XPATH, gC.POZISYON)
        experiences_page_title2= self.driver.find_element(By.XPATH, gC.SEKTOR)
        experiences_page_title3= self.driver.find_element(By.XPATH, gC.SEHİR_SECİNİZ)
        experiences_page_title4= self.driver.find_element(By.XPATH, gC.İS_BASLANGICI)
        experiences_page_title5= self.driver.find_element(By.XPATH, gC.İS_BİTİS)
        experiences_page_title6= self.driver.find_element(By.XPATH, gC.İS_ACIKLAMASI)
        assert experiences_page_title.text=="Kurum Adı*"
        assert experiences_page_title1.text=="Pozisyon*"
        assert experiences_page_title2.text=="Sektör*"
        assert experiences_page_title3.text=="Şehir Seçiniz*"
        assert experiences_page_title4.text=="İş Başlangıcı*"
        assert experiences_page_title5.text=="İş Bitiş*"
        assert experiences_page_title6.text=="İş Açıklaması"

#TEST CASE 2: “Doldurulması Zorunlu Alan” Uyarısı
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_requiredField(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)


        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        kaydet_button = self.driver.find_element(By.XPATH, gC.kaydet_button)
        kaydet_button.click()
        
        text_danger= self.driver.find_element(By.CLASS_NAME, gC.ZORUNLU_ALAN)
        assert text_danger.text=="Doldurulması zorunlu alan*"
        sleep(2)
        
#TEST CASE 3: Kurum Adı Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_company_name(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)


        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        text_kampus365 = self.driver.find_element(By.XPATH, "//input[@placeholder='Kampüs 365']")
        assert text_kampus365.get_attribute("placeholder") == "Kampüs 365"
        sleep(2)

        box1= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOX2_CASE3)))
        box1.send_keys(lessthanfive)

        kaydet_button = self.driver.find_element(By.XPATH, gC.kaydet_button)
        kaydet_button.click()

        lessthanFive= self.driver.find_element(By.CLASS_NAME, "text-danger")
        assert lessthanFive.text == "En az 5 karakter girmelisiniz"
        sleep(2)

        box1= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOX1_CASE3)))
        box1.send_keys(morethanfifty)

        morethanfifty= self.driver.find_element(By.XPATH,gC.MORE_THAN_FIFTY_CASE3 )
        assert morethanfifty.text == "En fazla 50 karakter girebilirsiniz"
        sleep(5)

#TEST CASE 4: Pozisyon Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_position(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        text_frontEndDeveloper = self.driver.find_element(By.XPATH, "//input[@placeholder='Front-End Developer']")
        assert text_frontEndDeveloper.get_attribute("placeholder") == "Front-End Developer"
        sleep(2)

        box_position= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOX_POSİTİON1)))
        box_position.send_keys(lessthanfive)

        kaydet_button = self.driver.find_element(By.XPATH, gC.kaydet_button)
        kaydet_button.click()

        lessthanFive= self.driver.find_element(By.XPATH, gC.LESS_THAN_FIVE_CASE4)
        assert lessthanFive.text == "En az 5 karakter girmelisiniz"
        sleep(2)

        box_position= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOX_POSİTİON)))
        box_position.send_keys(morethanfifty)

        morethanfifty= self.driver.find_element(By.XPATH, gC.MORE_THAN_FIFTY_CASE4 )
        assert morethanfifty.text == "En fazla 50 karakter girebilirsiniz"
        sleep(5)

#TEST CASE 5: Sektör Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_sector(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        text_yazılım = self.driver.find_element(By.XPATH, "//input[@placeholder='Yazılım']")
        assert text_yazılım.get_attribute("placeholder") == "Yazılım"
        sleep(2)

        box_sector= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOXSECTOR)))
        box_sector.send_keys(lessthanfive)

        kaydet_button = self.driver.find_element(By.XPATH, gC.kaydet_button)
        kaydet_button.click()

        lessthanFive= self.driver.find_element(By.XPATH, gC.LESS_THAN_FIVE)
        assert lessthanFive.text == "En az 5 karakter girmelisiniz"
        sleep(2)

        box_sector= WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.BOX_SECTOR)))
        box_sector.send_keys(morethanfifty)

        morethanfifty= self.driver.find_element(By.XPATH,gC.MORE_THAN_FIFTY )
        assert morethanfifty.text == "En fazla 50 karakter girebilirsiniz"
        sleep(5)

#TEST CASE 6: Şehir Seçiniz Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_city(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)

        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)   

        city = self.driver.find_element(By.XPATH,gC.CITY)
        city.click()
        
        text_city= self.driver.find_element(By.XPATH, gC.IL_SECINIZ)
        assert text_city.text == "İl Seçiniz"
        sleep(2)

#TEST CASE 7: İş Başlangıcı 
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_business_start(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(3)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        text_ggaayyyy = self.driver.find_element(By.XPATH, "//input[@placeholder='gg.aa.yyyy']")
        assert text_ggaayyyy.get_attribute("placeholder") == "gg.aa.yyyy"
        sleep(2)

        text_ggaayyyy = self.driver.find_element(By.XPATH, gC.TEXTT_GGAAYYYY)
        text_ggaayyyy.click()
        sleep(3)

        select_mounth= self.driver.find_element(By. CLASS_NAME,gC.SELECT_MOUNTH)
        select_mounth.click()
        sleep(2) 
        
        option_ocak = self.driver.find_element(By.XPATH, gC.OPTION_OCAK)
        option_ocak.click()
        sleep(2)

        select_year = self.driver.find_element(By.CLASS_NAME,gC.SELECT_YEAR)
        select_year.click()
        sleep(2)

        option_year = self.driver.find_element(By.XPATH, gC.OPTION_YEAR)
        option_year.click()
        sleep(2)

        day_element = self.driver.find_element(By.CSS_SELECTOR, gC.DAY_ELEMENT)
        day_element.click()
        sleep(2)
#TEST CASE 8: İş Bitiş 
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_businessFinish(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        sleep(2)

        loginbutton = self.driver.find_element(By.XPATH,gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

    
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)   

        text_ggaayyyy = self.driver.find_element(By.XPATH, gC.TEXTT_GGAAYYYY)
        text_ggaayyyy.click()
        sleep(3)

        select_mounth= self.driver.find_element(By. CLASS_NAME,gC.SELECT_MOUNTH)
        select_mounth.click()
        sleep(2) 
        
        option_ocak = self.driver.find_element(By.XPATH, gC.OPTION_OCAK)
        option_ocak.click()
        sleep(2)

        select_year = self.driver.find_element(By.CLASS_NAME,gC.SELECT_YEAR)
        select_year.click()
        sleep(2)

        option_year = self.driver.find_element(By.XPATH, gC.OPTION_YEAR)
        option_year.click()
        sleep(2)

        day_element = self.driver.find_element(By.CSS_SELECTOR, gC.DAY_ELEMENT)
        day_element.click()
        sleep(2)

        text_finisggaayyyy = self.driver.find_element(By.XPATH, gC.TEXTT_FINISH_GGAAYYYY)
        text_finisggaayyyy.click()
        sleep(3)

        select_mounth_finish= self.driver.find_element(By.XPATH,gC.SELECT_MOUNTH_FINISH)
        select_mounth_finish.click()
        sleep(2)
        
        option_ocak_finish = self.driver.find_element(By.XPATH, gC.OPTION_OCAK_FINISH) 
        option_ocak_finish.click()
        sleep(2)

        select_year_finish = self.driver.find_element(By.CLASS_NAME,gC.SELECT_YEAR_FINISH)
        select_year_finish.click()
        sleep(2)

        option_year_finish = self.driver.find_element(By.XPATH, gC.OPTION_YEAR_FINISH)
        option_year_finish.click()
        sleep(3)

        day_element_finish = self.driver.find_element(By.CSS_SELECTOR, gC.DAY_ELEMENT_FINISH)
        day_element_finish.click()
        sleep(2)
        
        continue_work_checkbox = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, gC.CONTİNUE_WORK_CHECKBOX)))
        continue_work_checkbox.click()
        sleep(2)
#TEST CASE 9: İş Açıklaması 
    @pytest.mark.parametrize("email, password, lessthanfive, morethanfifty, finishStartcalendarchoice, finishEndcalendarchoice, morethanthreehundered", getData())
    def test_businessDescription(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        sleep(2)

        loginbutton = self.driver.find_element(By.XPATH,gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)

    
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(2)

        experiences = self.driver.find_element(By.XPATH, gC.EXPERINCE_XPATH)
        experiences.click()
        sleep(2)

        description = self.driver.find_element(By. XPATH, gC.DESCRIPTION_BUSİNESS)
        description.click()
        description.send_keys(morethanthreehundered)
        sleep(2)

        save_button = self.driver.find_element(By.XPATH,gC.SAVE_BUTTON)
        save_button.click()
        sleep(2)

        self.driver.execute_script("window.scrollTo(0,500)")

        ucyuzkarakter = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.MORETHANTHREEHUNDERED)))
        assert ucyuzkarakter.text == "En fazla 300 karakter girebilirsiniz"

""" BUG: Gereksinimde iş bitiş seçim yılları sadece 2023-2024 olmalıdır, yazıyor. 
İş başlangıç tarihine göre iş bitiş tarihleri de değişmektedir. 
Buna göre 1974-2024 yılları arasını kapsıyor.  """


