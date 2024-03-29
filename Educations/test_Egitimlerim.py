from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl

from constants import egitimlerimConstans as c

class Test_Egitimlerim:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()
        sleep(3)  # import time
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
           self.driver.quit()    
    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data
        
    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_TumEgitimlerim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3)
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_DevamEgitimlerim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3)
        devamEttigimEgitimler=self.driver.find_element(By.XPATH,c.DEVAM_ETTIGIMEGITIMLER)
        devamEttigimEgitimler.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_TamaladıgımEgitimlerim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3)
        tamaladigimEgitimler=self.driver.find_element(By.XPATH,c.TAMAMLADIGIM_EGITIMLER)
        tamaladigimEgitimler.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)    

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_AramaKutusu(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        aramaKutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ARAMA_KUTUSU)))
        aramaKutusu.send_keys("Her")
        sleep(5) 
        self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/div[2]/div[1]/div/div[1]/div/button/div").click()
        sleep(3)  

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_AramaKutusuHatalıMetin(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        aramaKutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ARAMA_KUTUSU)))
        aramaKutusu.send_keys("h")
        sleep(5) 
        aramaKutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ARAMA_KUTUSU)))
        aramaKutusu.send_keys("e")
        sleep(5)  
        aramaKutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ARAMA_KUTUSU)))
        aramaKutusu.send_keys("r")
        sleep(5)  
        aramaKutusu = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ARAMA_KUTUSU)))
        aramaKutusu.send_keys("b")
        sleep(5)   
        self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/div[2]/div[1]/div/div[1]/div/button/div").click()
        sleep(3)  
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='all-lessons-tab-pane']/div/p")))
        assert toast_message.text == "Size atanan herhangi bir eğitim bulunmamaktadır." 
        sleep(3)
        devamEttigimEgitimler=self.driver.find_element(By.XPATH,c.DEVAM_ETTIGIMEGITIMLER)
        devamEttigimEgitimler.click()
        sleep(3) 
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='started-tab-pane']/div/p")))
        assert toast_message.text == "Devam ettiğiniz eğitim bulunmamaktadır."                        
        sleep(3)
        tamaladigimEgitimler=self.driver.find_element(By.XPATH,c.TAMAMLADIGIM_EGITIMLER)
        tamaladigimEgitimler.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH, "//*[@id='done-lessons-tab-pane']/div/p")))
        assert toast_message.text == "Tamamladığınız eğitim bulunmamaktadır." 
        sleep(3)
    
    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_KurumSecim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        kurumSecimiTik=self.driver.find_element(By.XPATH,c.KURUM_SECIMI_TIK)
        kurumSecimiTik.click()
        sleep(3)
        kurumSec=self.driver.find_element(By.XPATH,"//*[@id='react-select-2-listbox']")
        kurumSec.click()
        sleep(3)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_SeciliKurumKaldirma(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        kurumSecimiTik=self.driver.find_element(By.XPATH,c.KURUM_SECIMI_TIK)
        kurumSecimiTik.click()
        sleep(3)
        kurumSec=self.driver.find_element(By.XPATH,"//*[@id='react-select-2-listbox']")
        kurumSec.click()
        sleep(3)
        seciliKurumKaldirma=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/div[2]/div[1]/div/div[2]/div/div[1]/div[2]/div[1]")
        seciliKurumKaldirma.click()
        sleep(3)


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_KurumSecInput(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        kurumSecimiInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='react-select-2-input']")))
        kurumSecimiInput.send_keys("c")
        sleep(3)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_SiralamaSecim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        siralamaTik=self.driver.find_element(By.XPATH,c.SIRILAMA_SECIM_TIK)
        siralamaTik.click()
        sleep(3)
        siralamaSecim=self.driver.find_element(By.XPATH,"//div[@id='react-select-3-option-1']")
        siralamaSecim.click()
        sleep(5)
        self.driver.execute_script("window.scrollTo(0,900)")
        sleep(5)   

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_SiralamaSecimKaldirma(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        siralamaTik=self.driver.find_element(By.XPATH,c.SIRILAMA_SECIM_TIK)
        siralamaTik.click()
        sleep(3)
        siralamaSecim=self.driver.find_element(By.XPATH,"//div[@id='react-select-3-option-1']")
        siralamaSecim.click()
        sleep(5)  
        siralamaKaldirmaIkon=self.driver.find_element(By.XPATH,c.FILTRE_KALDIR_IKON)
        siralamaKaldirmaIkon.click()
        sleep(3)  

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_GirilenMetinSiralama(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(3) 
        girilenInputSiralama=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.GIRILEN_METINILE_ARAMA)))
        girilenInputSiralama.send_keys("Ü")
        sleep(5)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_EgitimeGit(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.EGITIMLERIM_URL)
        sleep(5) 
        self.driver.execute_script("window.scrollTo(0,300)")
        sleep(5)   
        egitimegit=self.driver.find_element(By.XPATH,c.EGITIME_GIT)
        egitimegit.click()
        sleep(15)
        


        