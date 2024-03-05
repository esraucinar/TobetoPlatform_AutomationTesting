from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl

from constants import degerlendirmeler as c

class Test_Degerlendirmeler:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()
        sleep(3)  # import time
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
           self.driver.quit()    
    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_degerlendirmeler(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        degerlendirmeler=self.driver.find_element(By.XPATH,c.DEGERLENDIRMELER)
        degerlendirmeler.click()
        sleep(3)
        baslaButon=self.driver.find_element(By.XPATH,c.BASLA_BUTON)
        baslaButon.click()
        sleep(2)


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_degerlendirmelerBackendTest(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        degerlendirmeler=self.driver.find_element(By.XPATH,c.DEGERLENDIRMELER)
        degerlendirmeler.click()
        sleep(3)   
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)
        backendTestButon=self.driver.find_element(By.XPATH,c.BACKEND_TEST_BUTON)
        backendTestButon.click()
        sleep(3) 


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_degerlendirmelerRaporGoruntule(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        degerlendirmeler=self.driver.find_element(By.XPATH,c.DEGERLENDIRMELER)
        degerlendirmeler.click()
        sleep(3)   
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3) 
        raporGoruntuleButon=self.driver.find_element(By.XPATH,c.RAPOR_TEST_BUTON)
        raporGoruntuleButon.click()
        sleep(5)
        raporGoruntule=self.driver.find_element(By.XPATH,c.RAPORGORUNTULE_BUTON)
        raporGoruntule.click()
        sleep(5)
        self.driver.find_element(By.XPATH,"//div/div/div/div/div/div[2]/button").click()
        sleep(3)



        

