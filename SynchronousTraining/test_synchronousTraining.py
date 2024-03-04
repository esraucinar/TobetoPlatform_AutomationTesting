from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import pytest

class Test_SynchronousTraining():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    def teardown_method(self): 
        self.driver.quit()
    def getData1():
        excel= openpyxl.load_workbook("data/emailPassword.xlsx")
        sheet= excel["Sayfa1"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            email=sheet.cell(i,1).value
            password= sheet.cell(i,2).value
            data.append((email,password))
        return data
#TEST CASE 1: “Eğitimi nasıl tamamlayabilirim?” ve “Eğitimi nasıl başarabilirim?” Ek Bilgilendirilmeleri
    @pytest.mark.parametrize("email,password", getData1())
    def test_ek_bilgilendirme(self,email,password):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        
        self.driver.execute_script("window.scrollTo(0,400)")
        sleep(2)

        egitimlerim = self.driver.find_element(By.XPATH,gC.LESSONS)
        egitimlerim.click()
        sleep(2)

        self.driver.execute_script("window.scrollTo(400,600)")
        sleep(5)
        show_more_button = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, gC.SHOWMORE_BUTTON)))
        show_more_button.click()
        sleep(2)
        self.driver.execute_script("window.scrollTo(0,2000)")
        sleep(3)
        egitime_git = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.PROCEED_TO_TRAINING)))
        egitime_git.click()
    
        tamamladın = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.YOU_DID_IT)))
        tamamladın.click()
        egitimi_nasıl_tamamlayabilirim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.HOW_CAN_I_COMPLETE_TRANING)))
        assert egitimi_nasıl_tamamlayabilirim.text == "Eğitimi nasıl tamamlayabilirim?"

        egitimi_nasıl_basarabilirim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.HOW_CAN_I_ACHIEVE_TRANING)))
        assert egitimi_nasıl_basarabilirim.text == "Eğitimi nasıl başarabilirim?"

#BUG: Video sayısı görünmüyor, kullanıcı 15 videonun hepsine katılmış bu oran da görünmüyor.
            
#TEST CASE 2: Kullanıcının Canlı Oturumlara Göre Aldığı Puanlama Sistemi ve  Beğeni- Favori İkonu İşlevleri
    @pytest.mark.parametrize("email,password", getData1())
    def test_canlıOturum_ikon(self,email,password):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        
        self.driver.execute_script("window.scrollTo(0,400)")
        sleep(2)

        egitimlerim = self.driver.find_element(By.XPATH,gC.LESSONS)
        egitimlerim.click()
        sleep(2)

        self.driver.execute_script("window.scrollTo(400,600)")
        sleep(5)
        show_more_button = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, gC.SHOWMORE_BUTTON)))
        show_more_button.click()
        sleep(2)
        self.driver.execute_script("window.scrollTo(0,2000)")
        sleep(3)
        egitime_git = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.PROCEED_TO_TRAINING)))
        egitime_git.click()
        sleep(3)

        like_icon= WebDriverWait(self.driver,30).until(ec.visibility_of_element_located((By.ID, gC.LIKE_ICON)))
        like_icon.click()
        sleep(2)

        like_number = self.driver.find_element(By.XPATH,gC.LIKE_NUMBER)
        like_number.click()
        sleep(2)

        close_window =WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.CLOSE_WINDOW)))
        close_window.click()
        sleep(3)

        like_number = self.driver.find_element(By.XPATH,gC.LIKE_NUMBER)
        like_number.click()

        user_click = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.USER_CLICK)))
        user_click.click()
        sleep(2)

        follow_she = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.FOLLOW_SHE)))
        follow_she.click()
        sleep(2)

        successful_follow_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, gC.SUCCESSFUL_FOLLOW_MESSAGE)))
        assert successful_follow_message.text == "Zehra Temizel isimli kullanıcıyı başarıyla takip ettin."
        sleep(2)

        close_message = WebDriverWait(self.driver,30).until(ec.visibility_of_element_located((By.XPATH,gC.CLOSE_MESSAGE)))
        close_message.click()
        sleep(2)
    
        close_window =WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.CLOSE_WINDOW)))
        close_window.click()
        sleep(3)

        like_number = self.driver.find_element(By.XPATH,gC.LIKE_NUMBER)
        like_number.click()

        user_click = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.USER_CLICK)))
        user_click.click()
        sleep(2)

        dontfollow_she = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.DONT_FOLLOW_SHE)))
        dontfollow_she.click()

        unsuccessful_follow_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.UNSUCCESSFUL_FOLLOW_MESSAGE)))
        assert unsuccessful_follow_message.text == "Zehra Temizel isimli kullanıcıyı başarıyla takipten çıktın." 

        close_unsuccessful_follow_message =WebDriverWait(self.driver,30).until(ec.visibility_of_element_located((By.XPATH,gC.CLOSE_UNSUCCESSFUL_FOLLOW_MESSAGE)))
        close_unsuccessful_follow_message.click()

        close_window =WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.CLOSE_WINDOW)))
        close_window.click()
        sleep(3)

        favori_icon = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.XPATH, gC.FAVORI_ICON)))
        favori_icon.click()

        favori_icon_message = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.FAVORI_ICON_MESSAGE)))
        assert favori_icon_message.text =="Favorilere ekleme işlemin başarıyla gerçekleşti."
        
#TEST CASE 3: İlerleme Çubuğunun İşlevi
    @pytest.mark.parametrize("email,password", getData1())
    def test_ilerleme_bar(self,email,password):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        
        self.driver.execute_script("window.scrollTo(0,400)")
        sleep(2)

        egitimlerim = self.driver.find_element(By.XPATH,gC.LESSONS)
        egitimlerim.click()
        sleep(2)

        self.driver.execute_script("window.scrollTo(400,600)")
        sleep(5)
        show_more_button = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH, gC.SHOWMORE_BUTTON)))
        show_more_button.click()
        sleep(2)
        self.driver.execute_script("window.scrollTo(0,2000)")
        sleep(3)
        egitime_git = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.PROCEED_TO_TRAINING)))
        egitime_git.click()
        sleep(3)   

        ilerleme_bar = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='dynamicContent']/div/div[1]/div/div[2]/div/div[2]/div/div/div/div/div")))
        ilerleme_bar.click()

        egitimi_tamamlama_oranı = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='body']/div[4]/div/div[2]/div")))
        assert egitimi_tamamlama_oranı.text == "Eğitim Tamamlama Oranı"

#“Eğitim Tamamlama Oranı” yazısı görülmelidir. Bu mesajla birlikte, kullanıcının tamamlama oranına göre, ilerleme çubuğunda renk değişimi ve oran gösterilmelidir.
#BUG: Kullanıcı videoları izlediği halde, böyle bir durum söz konusu değildir.    
        
        

      


        