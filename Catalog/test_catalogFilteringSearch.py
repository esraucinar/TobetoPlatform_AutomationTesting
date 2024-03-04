from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest

class Test_CatalogFilteringSearch():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    def teardown_method(self): 
        self.driver.quit()
    def getData2():
        excel= openpyxl.load_workbook("data/catalog.xlsx")
        sheet= excel["page1"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            start_learning=sheet.cell(i,1).value
            not_content =sheet.cell(i,2).value
            email=sheet.cell(i,3).value
            password= sheet.cell(i,4).value
            filtering = sheet.cell(i,5).value
            filtering1 = sheet.cell(i,6).value
            level =sheet.cell(i,7).value
            subject = sheet.cell(i,8).value
            software_language = sheet.cell(i,9).value
            instructor = sheet.cell (i,10).value
            situation =sheet.cell(i,11).value
            data.append((start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation))
        return data
#TEST CASE 1: Katalog Bölümü Arama Çubuğu İle Derslerin Listelenmesi
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_katalogBolumu_ders_listesi(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        search_bar = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.ID, gC.SEARCH)))
        search_bar.click()
        search_bar.send_keys(start_learning)
        sleep(3)
        magnifying_glass = self.driver.find_element(By.CSS_SELECTOR, gC.MAGNİFYİNG_GLASS)
        magnifying_glass.click()
        sleep(3)
        lesson = self.driver.find_element(By.CSS_SELECTOR, gC.LESSON)
        assert lesson.text == "Kısa Kısa: Kişisel Mükemmellik Seti"

#TEST CASE 2: "Aradığınız kriterlere uygun içerik bulunamadı." Uyarısının Kontrolü
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_uygunicerik_bulunamadı(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()  
        search_bar = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.ID, gC.SEARCH)))
        search_bar.click()
        search_bar.send_keys(not_content)
        sleep(3)  
        magnifying_glass = self.driver.find_element(By.CSS_SELECTOR, gC.MAGNİFYİNG_GLASS)
        magnifying_glass.click()
        sleep(3)
        assert self.driver.find_element(By.CSS_SELECTOR, gC.MESSAGE).text == "Aradığınız kriterlere uygun içerik bulunamadı."
#TEST CASE 3: Filtrele Paneli
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtrelePaneli(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click() 
        panel =WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.PANEL)))
        panel.click()

#TEST CASE 4: “Bana Özel” Penceresindeki Derslerin Listelenmesi
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_banaOzel(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()    
        bana_ozel = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,gC.SPECİAL_FOR_ME)))
        bana_ozel.click()
        sleep(2)
#TEST CASE 5: Filtrele Panelindeki Arama Kriteri
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_FiltrelePaneli_aramaKriteri(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        panel =WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.PANEL)))
        panel.click()
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)
        arama_kutusu = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.CSS_SELECTOR,gC.SEARCH_BOX)))
        arama_kutusu.click()
        arama_kutusu.send_keys(filtering)
        sleep(3)
        ucretsiz_egitimler = WebDriverWait(self.driver,20).until(ec.visibility_of_element_located((By.LINK_TEXT, gC.FREE_EDUCATION)))
        assert ucretsiz_egitimler.text == "Ücretsiz Eğitimler"
#TEST CASE 5:1
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleEgitimler(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()  
        egitimlerim =WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.EGİTİMLERİM)))
        egitimlerim.click()
        self.driver.execute_script("window.scrollTo(0,600)")
        arama_kutusu_egitimlerim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.ARAMA_KUTUSU_EGITIMLERIM)))
        arama_kutusu_egitimlerim.click()
        arama_kutusu_egitimlerim.send_keys(filtering1)
        sleep(3)
        dijital_gelisim = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, gC.DIJITAL_GELISIM)))
        assert dijital_gelisim.text=="Dijital Gelişim"
#TEST CASE 5:2
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleSeviye(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        sleep(5)
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)
        seviye = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.S_LEVEL)))
        seviye.click()
        arama_kutusu_seviye = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,gC.ARAMA_KUTUSU_SEVİYE)))
        arama_kutusu_seviye.click()
        arama_kutusu_seviye.send_keys(level)
        sleep(3)
        orta = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, "Orta")))
        assert orta.text=="Orta"
#TEST CASE 5:3
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleKonu(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        sleep(3) 
        self.driver.execute_script("window.scrollTo(0,600)")
        sleep(3)
        konu =  WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.KONU)))
        konu.click()
        self.driver.execute_script("window.scrollTo(600,700)")
        sleep(3)
        konu_aramaKutusu=WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.KONU_ARAMAKUTUSU)))
        konu_aramaKutusu.click()
        konu_aramaKutusu.send_keys(subject)
        sleep(3)
        esneklik = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, "Esneklik")))
        assert esneklik.text =="Esneklik"
#TEST CASE 5:4
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleYazılımDili(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        sleep(3) 
        self.driver.execute_script("window.scrollTo(0,600)")
        sleep(3)
        yazılım_dili = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.YAZILIM_DİLİ)))
        yazılım_dili.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(600,700)")
        yazılım_dili_aramaKutusu= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.YAZILIM_DİLİ_ARAMAKUTUSU)))
        yazılım_dili_aramaKutusu.click()
        yazılım_dili_aramaKutusu.send_keys(software_language)
        sleep(3)
        bootsrap= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, "Bootsrap")))
        assert bootsrap.text == "Bootsrap"
#TEST CASE 5:5
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleEgitmen(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        sleep(3) 
        self.driver.execute_script("window.scrollTo(0,600)")
        sleep(3) 
        egitmen = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.EGITMEN)))
        egitmen.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(600,700)")
        egitmen_aramaKutusu =WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.EGITMEN_ARAMAKUTUSU)))
        egitmen_aramaKutusu.click()
        egitmen_aramaKutusu.send_keys(instructor)
        sleep(3)
        egitmen_text= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, "İrem Balcı")))
        assert egitmen_text.text=="İrem Balcı"
#TEST CASE 5:6
    @pytest.mark.parametrize("start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation", getData2())
    def test_filtreleDurum(self,start_learning,not_content,email,password,filtering,filtering1,level,subject,software_language,instructor,situation):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        catalog = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.XPATH,gC.CATALOG)))
        catalog.click()
        sleep(3) 
        self.driver.execute_script("window.scrollTo(0,600)")
        sleep(3)
        durum = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, gC.DURUM)))
        durum.click()
        sleep(3)
        self.driver.execute_script("window.scrollTo(600,800)")
        durum_aramaKutusu= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR,gC.DURUM_ARAMAKUTUSU)))
        durum_aramaKutusu.click()
        durum_aramaKutusu.send_keys(situation)
        sleep(3)
        durum_text= WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.LINK_TEXT, "Henüz Başlanmamış Eğitimler")))
        assert durum_text.text =="Henüz Başlanmamış Eğitimler"
#BUG: İçeriklere tekrar tekrar tıkladığında içerikler değişiyor.
   



        
            
          
        

          
        

        

        