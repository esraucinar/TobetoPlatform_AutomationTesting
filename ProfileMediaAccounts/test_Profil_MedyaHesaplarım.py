from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl

from constants import profilMedyaHesaplarım as c

class Test_MedyaHesaplarim:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()
        sleep(3)  # import time
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
           self.driver.quit()    
    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_medyaHesaplarim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        medyaHesaplarim=self.driver.find_element(By.XPATH,c.MEDYA_HESAPLARIM)
        medyaHesaplarim.click()
        sleep(5)


        hesapSecme=self.driver.find_element(By.XPATH,c.MEDYA_HESAP_SECIM)
        hesapSecme.click()
        sleep(3)
        gitHubSecim=self.driver.find_element(By.XPATH,c.GITHUB_SECIM)
        gitHubSecim.click()
        sleep(3)
        gitHubAdres= WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.GITHUB_ADRES)))
        gitHubAdres.send_keys("https://github.com/yasinTrgt")
        kaydetButon=self.driver.find_element(By.XPATH,c.KAYDET_BUTON)
        kaydetButon.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Sosyal medya adresiniz başarıyla eklendi." 
        sleep(3)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_medyaHesaplarimGuncelleme(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(8)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        medyaHesaplarim=self.driver.find_element(By.XPATH,c.MEDYA_HESAPLARIM)
        medyaHesaplarim.click()
        sleep(5)
        medyaHesapGuncelleme=self.driver.find_element(By.XPATH,c.GUNCELLE_BUTONU)
        medyaHesapGuncelleme.click()
        sleep(5)
        self.driver.find_element(By.CSS_SELECTOR, ".modal-footer .form-control").click()
        guncellemeMetni=self.driver.find_element(By.CSS_SELECTOR,c.GUNCELLENECEK_METIN)
        guncellemeMetni.click()
        sleep(5)
        guncellemeGrilenMetin=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,c.GUNCELLEME_GIRILEN_METIN)))
        guncellemeGrilenMetin.send_keys("yeni")
        sleep(5)
        guncelleButon=self.driver.find_element(By.XPATH,c.GUNCELLE_BUTON)
        guncelleButon.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH,c.FORBIDDEN)))
        assert toast_message.text == "TOBETO\n• Forbidden" 
        sleep(3)


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_medyaHesaplarimSilme(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(8)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        medyaHesaplarim=self.driver.find_element(By.XPATH,c.MEDYA_HESAPLARIM)
        medyaHesaplarim.click()
        sleep(5)
        medyaHesapSil=self.driver.find_element(By.XPATH,c.MEDYA_HESAP_SIL_BUTON)
        medyaHesapSil.click()
        sleep(3)
        silEvetButon=self.driver.find_element(By.XPATH,c.SIL_EVET_BUTON)
        silEvetButon.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Sosyal medya adresiniz başarıyla kaldırıldı." 
        sleep(3)


        
        
  




