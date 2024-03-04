from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager . chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest

class Test_tobeto_giris():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    
    def teardown_method(self): 
        self.driver.quit() 

    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data
    
    def getDataUnsuccessful():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["unsuccessfulLogin1"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre=şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data
    def getDataUnsuccessful1():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["unsuccessfulLogin2"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre=şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data
    
    #successfulLogin
    @pytest.mark.parametrize("eposta,şifre", getData())
    def test_successfulLogin(self,eposta,şifre):
        epostainput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        epostainput.send_keys(eposta)
        sleep(3)

        password = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.NAME, gC.password)))
        password.send_keys(şifre)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)
        mesaj = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.CSS_SELECTOR,"div[class='toast-body']")))
        assert mesaj.text=="• Giriş başarılı."

    
    #unsuccessfulLogin1: Kayıtlı e-posta/şifre yanlış
    @pytest.mark.parametrize("eposta,şifre", getDataUnsuccessful())
    def test_unsuccessfulLogin(self,eposta,şifre):
        epostainput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        epostainput.send_keys(eposta)
        sleep(3)

        password = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.NAME, gC.password)))
        password.send_keys(şifre)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)
        assert self.driver.find_element(By.CSS_SELECTOR, ".toast-body").text == "• Geçersiz e-posta veya şifre."
    

    #unsuccessfulLogin2: Kayıtlı olmayan e-posta ile başarısız giriş.
    @pytest.mark.parametrize("eposta,şifre", getDataUnsuccessful1())
    def test_unsuccessfulLogin1(self,eposta,şifre):
        epostainput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        epostainput.send_keys(eposta)
        sleep(3)

        password = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.NAME, gC.password)))
        password.send_keys(şifre)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)
        assert self.driver.find_element(By.CSS_SELECTOR, ".toast-body").text == "• Henüz e-posta adresinizi doğrulamadınız."
    

    #Doldurulması zorunlu alan:e-posta yazıp şifre alanının boş kalmaması gerektiğini söyleyen başarısız giriş.
    @pytest.mark.parametrize("eposta,şifre", getDataUnsuccessful())
    def test_required_filed(self,eposta,şifre):
        
        epostainput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        epostainput.send_keys(eposta)
        sleep(3)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)   
        assert self.driver.find_element(By. XPATH, "//*[@id='__next']/div/main/section/div/div/div[1]/div/form/p").text == "Doldurulması zorunlu alan*"
    
    #Doldurulması zorunlu alan:e-posta alanını boş bıraktığında"doldurulması zorunlu alan"uyarısı, şifre yazılıp başarısız giriş. 
    @pytest.mark.parametrize("eposta,şifre", getDataUnsuccessful())
    def test_required_filed1(self,eposta,şifre):
        
        password = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.NAME, gC.password)))
        password.send_keys(şifre)

        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(4)  
        assert self.driver.find_element(By. XPATH, "//*[@id='__next']/div/main/section/div/div/div[1]/div/form/p")
    
        