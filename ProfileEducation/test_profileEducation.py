from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec 
from time import sleep
from constants import globalConstants as gC
import openpyxl
import pytest

class Test_profile_myEducation():
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(gC.URL)
        self.driver.maximize_window()
    def teardown_method(self): 
        self.driver.quit()
    def getData():
        excel= openpyxl.load_workbook("data/data.xlsx")
        sheet= excel["profile"]
        rows = sheet.max_row
        data=[]
        for i in range (2,rows+1):
            email=sheet.cell(i,1).value
            password= sheet.cell(i,2).value
            lessthanfive = sheet.cell(i,3).value
            morethanfifty = sheet.cell(i,4).value
            startcalendarchoice =sheet.cell(i,5).value
            finishStartcalendarchoice =sheet.cell(i,6).value
            finishEndcalendarchoice =sheet.cell(i,7).value
            morethanthreehundered =sheet.cell(i,8).value
            university = sheet.cell(i,9).value
            department = sheet.cell(i,10).value
            data.append((email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department))
        return data
#TEST CASE 1: “Eğitim Hayatım” Bölümündeki Alanların Listelenmesi   
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_myEducational_Life(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)
    
        myEducational_life_page_title= self.driver.find_element(By.XPATH, gC.EGITIMDURUMU)
        myEducational_life_page_title1= self.driver.find_element(By.XPATH,gC.UNIVERSITY)
        myEducational_life_page_title2= self.driver.find_element(By.XPATH,gC.MEZUNIYETYILI)
        myEducational_life_page_title3= self.driver.find_element(By.XPATH,gC.BOLUM)
        myEducational_life_page_title4= self.driver.find_element(By.XPATH,gC.BASLANGICYILI)
        assert myEducational_life_page_title.text=="Eğitim Durumu*"
        sleep(3)
        assert myEducational_life_page_title1.text=="Üniversite*"
        assert myEducational_life_page_title2.text=="Mezuniyet Yılı*"
        assert myEducational_life_page_title3.text=="Bölüm*"
        assert myEducational_life_page_title4.text=="Başlangıç Yılı*"
        sleep(2)

        #BUG: Gereksinimlerde belirtilen sektör başlığı listelenmiyor. işletim sistemi, macOS
#TEST CASE 2: “Doldurulması Zorunlu Alan” Uyarısı
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_field_mandatory(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)

        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()

        title_field_mandatory = self.driver.find_element(By.CLASS_NAME, gC.ZORUNLU_ALAN)
        assert title_field_mandatory.text== "Doldurulması zorunlu alan*" 
        sleep(2)

#TEST CASE 3: Üniversite Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_universityTitle(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3) 
        educational_status = self.driver.find_element(By.XPATH, gC.EDUCATIONAL_STATUS)
        educational_status.click()

        select_educational_status = self.driver.find_element(By.XPATH, gC.SELECT_EDUCATIONAL_STATUS)
        select_educational_status.click()

        university_field = self.driver.find_element(By.XPATH, gC. UNIVERSITY_FIELD)
        text_university_kampus365 = self.driver.find_element(By.XPATH, "//input[@placeholder='Kampüs 365']")
        assert text_university_kampus365.get_attribute("placeholder") == "Kampüs 365"
        sleep(2)
        university_field.click()
        university_field.send_keys("e")
        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()
        university_field_error = self.driver.find_element(By.XPATH,gC.ENAZIKIKARAKTER)
        assert university_field_error.text=="En az 2 karakter girmelisiniz"
        sleep(2)
        university_field.click()
        university_field.send_keys(morethanfifty)
        university_morethanfifty= self.driver.find_element(By.XPATH,gC.UMTFİFTY)
        sleep(3)
        assert university_morethanfifty.text == "En fazla 50 karakter girebilirsiniz"
        sleep(3)
#TEST CASE 4: Bölüm Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_departmentTitle(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)

        department_title_yazılım = self.driver.find_element(By.XPATH, "//input[@placeholder='Yazılım']")
        assert department_title_yazılım.get_attribute("placeholder")=="Yazılım"
        sleep(2)
        department_title_yazılım.send_keys("a")
        sleep(2)
        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()
        sleep(2)
        department_field_error = self.driver.find_element(By.XPATH,gC.BOLUM_ENAZIKIKARAKTER)
        assert department_field_error.text=="En az 2 karakter girmelisiniz"
        sleep(2)
        
        department_title_yazılım.send_keys(morethanfifty)
        department_morethanfifty= self.driver.find_element(By.XPATH,gC.BOLUM_ENFAZLAELLIKARAKTER)
        assert department_morethanfifty.text == "En fazla 50 karakter girebilirsiniz"
        sleep(3)
#TEST CASE 5: Başlangıç Yılı Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_departmentTitle(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)

        start_year_default = self.driver.find_element(By.XPATH,"//input[@placeholder='Başlangıç Yılınızı Seçiniz']") 
        assert start_year_default.get_attribute("placeholder")=="Başlangıç Yılınızı Seçiniz"
        sleep(2)
        start_year= self.driver.find_element(By.XPATH,gC.START_YEAR)
        start_year.click()
        sleep(3)
        select_start_year= self.driver.find_element(By.CSS_SELECTOR, gC.SELECTSTARTYEAR)
        select_start_year.click()
        sleep(3)
#TEST CASE 6: Mezuniyet Yılı Başlığındaki Görünüm ve Uyarılar
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_graduationYear(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)

        start_year= self.driver.find_element(By.XPATH,gC.START_YEAR)
        start_year.click()
        sleep(3)
        select_start_year= self.driver.find_element(By.CSS_SELECTOR, gC.SELECTSTARTYEAR)
        select_start_year.click()
        sleep(3)

        graduation_year_default = self.driver.find_element(By.XPATH,"//input[@placeholder='Mezuniyet Yılınızı Seçiniz']") 
        assert graduation_year_default.get_attribute("placeholder")=="Mezuniyet Yılınızı Seçiniz"
        sleep(2)
        graduation_year = self.driver.find_element(By.XPATH,gC.GRADUATİON)
        graduation_year.click()
        sleep(3)

        for _ in range(4):
            next_button = WebDriverWait(self.driver, 10).until(ec.element_to_be_clickable((By.CSS_SELECTOR, gC.FOR)))
            next_button.click()

        select_graduation_year = self.driver.find_element(By.CSS_SELECTOR, gC.SELECT_GRADUATIONYEAR)
        select_graduation_year.click()
        sleep(2)
        graduation_year_checkbox = self.driver.find_element(By.XPATH, gC.GRADUATİON_YEAR_CHECKBOX)
        graduation_year_checkbox.click()
        sleep(2)

#TEST CASE 7: “Eğitim Bilgisi Eklendi” uyarısı 
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_education_information_added(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)

        educational_status = self.driver.find_element(By.XPATH, gC.EDUCATIONAL_STATUS)
        educational_status.click()
        select_educational_status = self.driver.find_element(By.XPATH, gC.SELECT_EDUCATIONAL_STATUS)
        select_educational_status.click()
        sleep(2)

        university_field = self.driver.find_element(By.XPATH, gC. UNIVERSITY_FIELD)
        university_field.click()
        university_field.send_keys(university)
        sleep(2)

        department_title_yazılım = self.driver.find_element(By.XPATH, "//input[@placeholder='Yazılım']")
        department_title_yazılım.click()
        department_title_yazılım.send_keys(department)
        sleep(2)

        start_year_default = self.driver.find_element(By.XPATH,"//input[@placeholder='Başlangıç Yılınızı Seçiniz']")
        start_year_default.click()
        self.driver.find_element(By.CSS_SELECTOR, gC.BACK).click()
        self.driver.find_element(By.CSS_SELECTOR, gC. START_2015).click()
        sleep(2)

        graduation_year = self.driver.find_element(By.XPATH,gC.GRADUATİON)
        graduation_year.click()
        sleep(3)

        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYET_GERİGELME).click()
        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYETYILI_SECME).click()
        sleep(2)

        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()
        toast_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Eğitim bilgisi eklendi."
        sleep(2)

#TEST CASE 8: “Seçilen Eğitimi Silmek İstediğinize Emin misiniz? Bu İşlem Geri Alınamaz” uyarısı ve Eğitim Bilgisini Silebilme İşlemi
    @pytest.mark.parametrize("email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department",getData())
    def test_education_delete(self,email,password,lessthanfive,morethanfifty,finishStartcalendarchoice,finishEndcalendarchoice,morethanthreehundered,university,department):
        emailinput = WebDriverWait(self.driver,15).until(ec.visibility_of_element_located((By.XPATH, gC.name)))
        emailinput.send_keys(email)
        sleep(2)
        password_input = self.driver.find_element(By.XPATH, gC.password_input)
        password_input.send_keys(password)
        loginbutton = self.driver.find_element(By.XPATH, gC.LOGGİNBUTTON_XPATH)
        loginbutton.click()
        sleep(5)
        profilebutton =self.driver.find_element(By.XPATH, gC.PROFILE_XPATH)
        profilebutton.click()
        sleep(3)
        informationprofile = self.driver.find_element(By.XPATH, gC.INFORMATİON_PROFILE)
        informationprofile.click()
        sleep(3)
        myEducational_life =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.MYEDUCATIONALİFE)))
        myEducational_life.click()
        sleep(3)   
        educational_status = self.driver.find_element(By.XPATH, gC.EDUCATIONAL_STATUS)
        educational_status.click()
        select_educational_status = self.driver.find_element(By.XPATH, gC.SELECT_EDUCATIONAL_STATUS)
        select_educational_status.click()
        sleep(2)
        university_field = self.driver.find_element(By.XPATH, gC. UNIVERSITY_FIELD)
        university_field.click()
        university_field.send_keys(university)
        sleep(2)
        department_title_yazılım = self.driver.find_element(By.XPATH, "//input[@placeholder='Yazılım']")
        department_title_yazılım.click()
        department_title_yazılım.send_keys(department)
        sleep(2)
        start_year_default = self.driver.find_element(By.XPATH,"//input[@placeholder='Başlangıç Yılınızı Seçiniz']")
        start_year_default.click()
        self.driver.find_element(By.CSS_SELECTOR, gC.BACK).click()
        self.driver.find_element(By.CSS_SELECTOR, gC. START_2015).click()
        sleep(2)
        graduation_year = self.driver.find_element(By.XPATH,gC.GRADUATİON)
        graduation_year.click()
        sleep(3)
        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYET_GERİGELME).click()
        self.driver.find_element(By.CSS_SELECTOR, gC.MEZUNIYETYILI_SECME).click()
        sleep(2)
        saveButton = self.driver.find_element(By.XPATH,gC.SAVEBUTTON)
        saveButton.click()
        self.driver.execute_script("window.scrollTo(0,500)")
        sleep(3)
        
        delete= WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.XPATH,gC.DELETE)))
        delete.click()
        alert_message = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,gC.ALERT_MESSAGE)))
        assert alert_message.text == "Seçilen eğitimi silmek istediğinize emin misiniz?"
        sleep(2)
        yes_button = self.driver.find_element(By.XPATH,gC.YES_BUTTON)
        yes_button.click()
        
        toast_message = WebDriverWait(self.driver,10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Eğitim kaldırıldı."
        sleep(3)

        


        















        








        
   
        