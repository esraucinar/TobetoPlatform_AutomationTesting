from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains
import pytest
import openpyxl

from constants import profilYabanciDillerim as c

class Test_Yabanci_Dillerim:
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()
        sleep(3)  # import time
    
    def teardown_method(self): # her testinin bitiminde çalışacak fonk
           self.driver.quit()    
    def getData():
        excel= openpyxl.load_workbook("data/giris.xlsx")
        sheet= excel["successfulLogin"]
        rows=sheet.max_row
        data=[]
        for i in range(2,rows+1):
            eposta=sheet.cell(i,1).value
            şifre= sheet.cell(i,2).value
            data.append((eposta,şifre))
        return data

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_loginAndProfilim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        yabanciDillerim=self.driver.find_element(By.XPATH,c.PROFIL_YABANCI_DILLERIM)
        yabanciDillerim.click()
        sleep(5)


        # profilim=self.driver.find_element(By.XPATH,c.PROFILIM)
        # profilim.click()
        # sleep(3)
        # profilDuzenle=self.driver.find_element(By.XPATH,c.PROFIL_DUZENLE_SIMGE)
        # profilDuzenle.click()
        # sleep(3)
    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_loginAndProfilimKaydet(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        yabanciDillerim=self.driver.find_element(By.XPATH,c.PROFIL_YABANCI_DILLERIM)
        yabanciDillerim.click()
        sleep(5) 
        kaydetButon=self.driver.find_element(By.XPATH,c.KAYDET_BUTTON)  
        kaydetButon.click()
        sleep(3)
        dolrulmasıZorunluAlan=self.driver.find_element(By.XPATH,c.DOLDURULMASI_ZORUNLU_ALAN_KONTROL)
        assert dolrulmasıZorunluAlan.text == c.DOLDURULMASI_ZORUNLU_ALAN_METNI
        

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_loginAndProfilimDilSecim(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        yabanciDillerim=self.driver.find_element(By.XPATH,c.PROFIL_YABANCI_DILLERIM)
        yabanciDillerim.click()
        sleep(5) 
        yabanciDilSecim=self.driver.find_element(By.XPATH,c.DIL_SECIMI)
        yabanciDilSecim.click()
        sleep(5)
        ingilizceSecim=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[1]/select/option[12]")
        ingilizceSecim.click()
        sleep(5)
        seviyeSecim=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[2]/select")
        seviyeSecim.click()
        sleep(5)
        seviyeSec=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[2]/select/option[3]")
        seviyeSec.click()
        sleep(3)
        kaydetButon=self.driver.find_element(By.XPATH,c.KAYDET_BUTTON)  
        kaydetButon.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Yabancı dil bilgisi eklendi." 
        sleep(3)


    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_loginAndProfilimAyniDilSecme(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        yabanciDillerim=self.driver.find_element(By.XPATH,c.PROFIL_YABANCI_DILLERIM)
        yabanciDillerim.click()
        sleep(5) 
        yabanciDilSecim=self.driver.find_element(By.XPATH,c.DIL_SECIMI)
        yabanciDilSecim.click()
        sleep(5)
        ingilizceSecim=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[1]/select/option[12]")
        ingilizceSecim.click()
        sleep(5)
        seviyeSecim=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[2]/select")
        seviyeSecim.click()
        sleep(5)
        seviyeSec=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/form/div/div[2]/select/option[3]")
        seviyeSec.click()
        sleep(3)
        kaydetButon=self.driver.find_element(By.XPATH,c.KAYDET_BUTTON)  
        kaydetButon.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Bu dil zaten mevcut." 
        sleep(3)

    @pytest.mark.parametrize("eposta,şifre",getData())
    def test_loginAndProfilimDilSilme(self,eposta,şifre):
        loginButton = self.driver.find_element(By.LINK_TEXT, "Giriş Yap")
        loginButton.click()
        sleep(3)
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.E_MAIL)))
        usernameInput.send_keys(eposta)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME,c.PASSWORD_NAME)))
        passwordInput.send_keys(şifre)
        loginButtonClick= self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_CLICK_XPATH)
        loginButtonClick.click()
        sleep(3)  
        self.driver.get(c.PROFIL_BILGILERIM_URL) 
        sleep(5)
        yabanciDillerim=self.driver.find_element(By.XPATH,c.PROFIL_YABANCI_DILLERIM)
        yabanciDillerim.click()
        sleep(5) 
        dilSecim=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/div/div/div/span[1]")
        dilSecim.click()
        sleep(3)
        dilSilmeButton=self.driver.find_element(By.XPATH,"//*[@id='__next']/div/main/section/div/div/div[2]/div/div/div/span[2]")
        dilSilmeButton.click()
        sleep(3)
        evetButton=self.driver.find_element(By.XPATH,"/html/body/div[3]/div/div/div/div/div/div[2]/button[2]")
        evetButton.click()
        sleep(3)
        toast_message = WebDriverWait(self.driver, 10).until(ec.visibility_of_element_located((By.CSS_SELECTOR, ".toast-body")))
        assert toast_message.text == "• Yabancı dil kaldırıldı." 
        sleep(3)






    
        
